import openpyxl

book=openpyxl.load_workbook("C:\\Users\\Adil\\Downloads\\Book1.xlsx")

sheet=book.active


cell=sheet.cell(row=2,column=3)

print(cell.value)

print(sheet.cell(row=2,column=4).value)

print(sheet.max_row)

print(sheet.max_column)
print("------------------------------------------------------")
data_dict = {}

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column + 1):
            key = sheet.cell(row=1, column=j).value    # header row
            value = sheet.cell(row=2, column=j).value  # data row
            data_dict[key] = value

print(data_dict)

#dictionary using


#loops working structure
#irst, i = 1
#→ then j = 1, j = 2, j = 3

#Next, i = 2
#→ again j = 1, j = 2, j = 3


#Outer loop i = 1 (Row 1)

#Inner loop: print A1, B1, C1, D1

#Outer loop i = 2 (Row 2)

#Inner loop: print A2, B2, C2, D2

#Outer loop i = 3 (Row 3)

#Inner loop: print A3, B3, C3, D3